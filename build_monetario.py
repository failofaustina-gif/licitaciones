#!/usr/bin/env python3
"""
Genera monetario.html (explorador interactivo) e informe.html (portada) a partir de
dos únicas fuentes oficiales:
  - series.xlsm (https://www.bcra.gob.ar/datos-monetarios-diarios/): indicadores monetarios.
  - colocaciones_deuda.xlsx (Oficina Nacional de Crédito Público): licitaciones del Tesoro.

Uso: python3 build_monetario.py [ruta_xlsm] [ruta_xlsx_colocaciones]
"""
import sys, re, json, datetime
import openpyxl
from pathlib import Path

HERE = Path(__file__).parent
XLSM = Path(sys.argv[1]) if len(sys.argv) > 1 else HERE / "series.xlsm"

MAX_POINTS = 1000  # ~4 años hábiles por serie, para no inflar el JSON

SHEETS = {
    "BASE MONETARIA": "BM",
    "RESERVAS": "RES",
    "DEPOSITOS": "DEP",
    "PRESTAMOS": "PRE",
    "TASAS DE MERCADO": "TAS",
    "INSTRUMENTOS DEL BCRA": "INS",
}

# ---------- 1. leer series.xlsm (todas las columnas) ----------
wb = openpyxl.load_workbook(XLSM, data_only=True, read_only=False, keep_vba=False)

def resolve_headers(ws, header_rows=(4, 5, 6, 7)):
    """Arma el texto de cada columna combinando las filas de encabezado,
    respetando las celdas combinadas (merge) reales de la hoja."""
    grid = {}
    for r in header_rows:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                grid[(r, c)] = str(v).replace("\n", " ").strip()
    for m in ws.merged_cells.ranges:
        if m.max_row < min(header_rows) or m.min_row > max(header_rows):
            continue
        top_val = ws.cell(row=m.min_row, column=m.min_col).value
        if top_val in (None, ""):
            continue
        top_val = str(top_val).replace("\n", " ").strip()
        for r in range(max(m.min_row, min(header_rows)), min(m.max_row, max(header_rows)) + 1):
            for c in range(m.min_col, m.max_col + 1):
                grid[(r, c)] = top_val
    labels = {}
    for c in range(2, ws.max_column + 1):
        parts = []
        for r in header_rows:
            v = grid.get((r, c))
            if v and (not parts or parts[-1] != v):
                parts.append(v)
        labels[c] = " - ".join(parts) if parts else f"Columna {c}"
    return labels

def classify_kind(sheet, label):
    low = label.lower()
    last_segment = label.split(" - ")[-1].strip().lower()
    if sheet == "RESERVAS" and "tipo de cambio" in low:
        return "fx"
    if sheet == "TASAS DE MERCADO" or last_segment in ("tna", "tea") or "tasa de política monetaria" in low or "tasa de interés" in low:
        return "rate"
    if sheet == "RESERVAS":
        return "usd"
    if "expresados en dólares" in low:
        return "usd"
    if "expresados en pesos" in low:
        return "ars"
    if "en dólares" in low or "total dólares" in low or "total \ndólares".replace("\n", " ") in low:
        return "usd"
    return "ars"

def read_sheet(sheet_name, prefix):
    ws = wb[sheet_name]
    labels = resolve_headers(ws)
    cols = [c for c in range(2, ws.max_column + 1) if labels.get(c) and "tipo de serie" not in labels[c].lower()]
    buffers = {c: [] for c in cols}
    last_date = None
    for row in ws.iter_rows(min_row=8, values_only=True):
        d = row[0]
        if not isinstance(d, (datetime.date, datetime.datetime)):
            continue
        if last_date is not None and d < last_date:
            break
        last_date = d
        ds = d.strftime("%Y-%m-%d")
        for c in cols:
            idx = c - 1
            v = row[idx] if idx < len(row) else None
            if isinstance(v, (int, float)):
                buffers[c].append((ds, round(v, 4)))
    out = []
    for c in cols:
        pts = buffers[c][-MAX_POINTS:]
        if not pts:
            continue
        out.append({
            "id": f"{prefix}_{c}",
            "sheet": sheet_name,
            "label": labels[c],
            "kind": classify_kind(sheet_name, labels[c]),
            "data": pts,
        })
    return out

all_series = []
for sheet_name, prefix in SHEETS.items():
    all_series.extend(read_sheet(sheet_name, prefix))

as_of = None
for s in all_series:
    if s["data"]:
        d = s["data"][-1][0]
        if as_of is None or d > as_of:
            as_of = d

HIST_DIR = HERE / "licitaciones_historial"

# ---------- 2a-bis. leer colocaciones_deuda.xlsx (registro de la Oficina Nacional de --------
# ---------- Crédito Público) y armar UNA licitación por cada fecha de colocación -----------
# Fuente mucho más robusta que datos_informe.xlsx/el PDF: es una tabla de columnas fijas
# (Nombre del Instrumento, Fecha colocación, Valor Nominal, Valor Efectivo, etc.) que no
# cambia de formato de una publicación a otra. No trae ofertas recibidas ni montos ofertados
# (eso solo lo tiene el comunicado en PDF de la Secretaría de Finanzas) — únicamente lo
# efectivamente adjudicado. Para actualizar: reemplazar colocaciones_deuda.xlsx por la
# versión nueva que publique la Oficina Nacional de Crédito Público y volver a correr esto
# (o simplemente pushear — el workflow lo hace solo).
XLSX_COLOC = Path(sys.argv[2]) if len(sys.argv) > 2 else HERE / "colocaciones_deuda.xlsx"

COLOC_SHEETS = {
    "Bonos": "bono", "Letras": "letra",
    "Otras Operaciones": "otra_operacion",
    # OJO: dos hojas quedan afuera a propósito.
    # - "Canjes- Conversiones": estructura totalmente distinta (dos sub-tablas apiladas,
    #   "Canje" y "Conversiones", esta última con columnas "Instrumento a entregar" /
    #   "Instrumento a dar de baja" en vez de "Valor Efectivo") — no hay un VE colocado
    #   comparable con las demás hojas, y tratar de encajarla en el mismo parseo termina
    #   leyendo nombres de instrumentos donde se esperan números.
    # - "Letras ISP" (intra-sector público): por definición son deuda entre organismos del
    #   Estado (p. ej. letras técnicas emitidas directamente al BCRA), no colocaciones de
    #   mercado — y de hecho mezclan magnitudes completamente distintas al resto (algunas
    #   filas de "LETRA/U$S/BCRA/..." aparecen con valores ~1000x más grandes que instrumentos
    #   comparables), así que no son comparables ni siquiera dentro de la propia hoja.
}
TIPO_LABELS_COMPOSICION = {
    "bono": "Bonos", "letra": "Letras", "letra_isp": "Letras ISP",
    "otra_operacion": "Otras operaciones", "canje": "Canjes/conversiones",
}

def excel_date_to_iso(v):
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        return (datetime.date(1899, 12, 30) + datetime.timedelta(days=v)).strftime("%Y-%m-%d")
    return None

def find_header_row(ws, max_row=10):
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() == "Nombre del Instrumento":
            return r
    return None

def parse_colocaciones_excel(path):
    wb3 = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sheet_name, tipo in COLOC_SHEETS.items():
        if sheet_name not in wb3.sheetnames:
            continue
        ws = wb3[sheet_name]
        header_row = find_header_row(ws)
        if header_row is None:
            continue
        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row, column=c).value
            headers.append(re.sub(r"\s+", " ", str(v)).strip() if v not in (None, "") else "")

        def col(name):
            for i, h in enumerate(headers):
                if h.startswith(name):
                    return i
            return -1

        c_nombre, c_emision, c_venc = col("Nombre del Instrumento"), col("Fecha de emisión"), col("Vencimiento")
        c_moneda, c_moneda_origen = col("Tipo Moneda"), col("Moneda de Origen")
        c_coloc, c_vn, c_ve = col("Fecha colocación"), col("Valor Nominal"), col("Valor Efectivo")
        c_vta, c_precio, c_vida = col("Valor Técnico Adjudicado"), col("Precio de emisión"), col("Vida Promedio")
        c_resol = col("Resolución") if col("Resolución") != -1 else col("Norma")
        c_numero = col("Número")

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if c_nombre == -1 or c_nombre >= len(row) or row[c_nombre] in (None, ""):
                continue
            fecha_colocacion = excel_date_to_iso(row[c_coloc]) if 0 <= c_coloc < len(row) else None
            if not fecha_colocacion:
                continue
            def get(idx):
                return row[idx] if 0 <= idx < len(row) else None
            def get_num(idx):
                # defensivo: si la columna esperada no es numérica (p. ej. el layout de la
                # hoja no coincide con lo esperado), se descarta en vez de romper el build.
                v = get(idx)
                return v if isinstance(v, (int, float)) else None
            out.append({
                "tipo": tipo, "hoja": sheet_name, "nombre": str(row[c_nombre]).strip(),
                "fecha_emision": excel_date_to_iso(get(c_emision)) if c_emision != -1 else None,
                "vencimiento": excel_date_to_iso(get(c_venc)) if c_venc != -1 else None,
                "moneda": get(c_moneda), "moneda_origen": get(c_moneda_origen),
                "fecha_colocacion": fecha_colocacion,
                "valor_nominal": get_num(c_vn), "precio_emision": get_num(c_precio), "vida_promedio": get_num(c_vida),
                "valor_efectivo": get_num(c_ve) if get_num(c_ve) is not None else get_num(c_vta),
                "resolucion": get(c_resol), "numero": get(c_numero),
            })
    return out

def es_instrumento_nuevo(nombre, fecha_colocacion, todas):
    return not any(c["nombre"] == nombre and c["fecha_colocacion"] < fecha_colocacion for c in todas)

def fmt_date_ddmmyyyy(iso):
    y, m, d = iso.split("-")
    return f"{d}/{m}/{y}"

def build_licitaciones_from_colocaciones(colocaciones):
    fechas = sorted({c["fecha_colocacion"] for c in colocaciones if c["fecha_colocacion"]})
    licitaciones_out = []
    for fecha in fechas:
        items = [dict(c, nuevo=es_instrumento_nuevo(c["nombre"], c["fecha_colocacion"], colocaciones))
                 for c in colocaciones if c["fecha_colocacion"] == fecha]
        ve_ars = sum((c["valor_efectivo"] or 0) for c in items if c["moneda"] == "MONEDA NACIONAL")
        ve_usd = sum((c["valor_efectivo"] or 0) for c in items if c["moneda"] == "MONEDA EXTRANJERA")
        by_tipo = {}
        for c in items:
            if c["moneda"] != "MONEDA NACIONAL" or not c["valor_efectivo"]:
                continue
            by_tipo[c["tipo"]] = by_tipo.get(c["tipo"], 0) + c["valor_efectivo"]
        composicion = [{"categoria": TIPO_LABELS_COMPOSICION.get(t, t), "monto": m, "pct": m / ve_ars}
                       for t, m in by_tipo.items()] if ve_ars else []
        licitaciones_out.append({
            "origen": "excel", "fecha": fecha,
            "titulo": f"Licitación del Tesoro — {fmt_date_ddmmyyyy(fecha)}",
            "subtitulo": f"Fuente: registro de Colocaciones de Deuda ({len(items)} instrumento{'s' if len(items)!=1 else ''} colocado{'s' if len(items)!=1 else ''}).",
            "resumenExcel": {"veAdjudicadoArs": ve_ars, "veAdjudicadoUsd": ve_usd, "cantidadInstrumentos": len(items)},
            "composicion": composicion, "instrumentosExcel": items,
        })
    return licitaciones_out

if XLSX_COLOC.exists():
    colocaciones_all = parse_colocaciones_excel(XLSX_COLOC)
    if colocaciones_all:
        nuevas_lic = build_licitaciones_from_colocaciones(colocaciones_all)
        HIST_DIR.mkdir(exist_ok=True)
        for lic in nuevas_lic:
            path = HIST_DIR / f"{lic['fecha']}.json"
            existing = {}
            if path.exists():
                try:
                    existing = json.loads(path.read_text(encoding="utf-8"))
                except (json.JSONDecodeError, OSError):
                    existing = {}
            merged = dict(existing)
            merged["fecha"] = lic["fecha"]
            merged["titulo"] = existing.get("titulo") or lic["titulo"]
            merged["subtitulo"] = existing.get("subtitulo") or lic["subtitulo"]
            merged["resumenExcel"] = lic["resumenExcel"]
            merged["instrumentosExcel"] = lic["instrumentosExcel"]
            merged["composicion"] = lic["composicion"]
            merged["origen"] = "excel"
            path.write_text(json.dumps(merged, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Colocaciones de deuda: {len(colocaciones_all)} filas -> {len(nuevas_lic)} licitaciones archivadas en licitaciones_historial/.")
    else:
        print(f"Aviso: {XLSX_COLOC.name} no tiene ninguna fila reconocible (¿hojas Bonos/Letras/etc. con formato distinto?).")

# ---------- 2b. cargar todo el historial de licitaciones archivadas ----------
licitaciones = []
if HIST_DIR.exists():
    for f in HIST_DIR.glob("*.json"):
        try:
            licitaciones.append(json.loads(f.read_text(encoding="utf-8")))
        except (json.JSONDecodeError, OSError) as e:
            print(f"Aviso: no se pudo leer {f.name} del historial de licitaciones: {e}")
licitaciones.sort(key=lambda x: x.get("fecha") or "", reverse=True)

# ---------- 2c. PBI nominal trimestral (para expresar series como % del PBI) ----------
# Se actualiza a mano en pbi_trimestral.json (INDEC publica con ~3 meses de rezago y no tiene
# un endpoint estable para automatizar la descarga). Los valores de ese archivo son los mismos
# que las columnas "I/II/III/IV trimestre" del Cuadro 8 de INDEC: cada uno YA está a tasa
# anualizada (no es el flujo real del trimestre), así que se usan directamente como denominador
# anual por trimestre — NO hay que sumar los 4 trimestres de un año (eso daría ~4x de más).
QTR_START = {"Q1": "01-01", "Q2": "04-01", "Q3": "07-01", "Q4": "10-01"}
pbi_quarters = []
pbi_path = HERE / "pbi_trimestral.json"
if pbi_path.exists():
    pbi_raw = json.loads(pbi_path.read_text(encoding="utf-8"))["trimestres"]
    for q in sorted(pbi_raw.keys()):
        year, qn = q.split("-Q")
        pbi_quarters.append({"trimestre": q, "desde": f"{year}-{QTR_START[f'Q{qn}']}", "anualizado": round(pbi_raw[q], 1)})
    pbi_quarters.sort(key=lambda x: x["desde"])
else:
    print("Aviso: no se encontró pbi_trimestral.json, no se va a poder expresar nada como % del PBI.")

# ---------- 2d. calendario de publicaciones (IPC, PBI trimestral, licitaciones) ----------
# Igual que pbi_trimestral.json: se actualiza a mano, no hay endpoint estable para ninguna de
# las tres fuentes. Ver instrucciones adentro del archivo.
calendario = {"eventos": [], "recurrentes": []}
calendario_path = HERE / "calendario.json"
if calendario_path.exists():
    calendario_raw = json.loads(calendario_path.read_text(encoding="utf-8"))
    calendario = {
        "eventos": sorted(calendario_raw.get("eventos", []), key=lambda e: e.get("fecha") or ""),
        "recurrentes": calendario_raw.get("recurrentes", []),
    }
else:
    print("Aviso: no se encontró calendario.json, la pestaña Calendario va a quedar vacía.")

dashboard_data = {
    "generated_at": datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
    "as_of": as_of,
    "series": all_series,
    "licitaciones": licitaciones,
    "pbi": pbi_quarters,
    "calendario": calendario,
}

json_str = json.dumps(dashboard_data, ensure_ascii=False, separators=(",", ":"))
print(f"Series exportadas: {len(all_series)} | as_of={as_of} | JSON: {len(json_str)/1024:.0f} KB")

# ---------- 3. armar monetario.html ----------
TEMPLATE = (HERE / "_monetario_template.html").read_text(encoding="utf-8")
html = TEMPLATE.replace("__DASHBOARD_DATA__", json_str).replace("__AS_OF__", as_of or "s/d")

out_html = HERE / "monetario.html"
out_html.write_text(html, encoding="utf-8")
print(f"HTML generado: {out_html}")

# ---------- 4. balance simplificado del BCRA (activo/pasivo), para la portada ----------
series_by_id = {s["id"]: s for s in all_series}
def last_val(sid):
    s = series_by_id.get(sid)
    return s["data"][-1][1] if s and s["data"] else None

res_usd, fx, bm, pases, letras = (last_val(i) for i in ("RES_3", "RES_16", "BM_30", "INS_2", "INS_7"))
activo_pasivo = None
if res_usd is not None and fx is not None and bm is not None:
    activo = res_usd * fx
    pasivo = bm + (pases or 0) + (letras or 0)
    activo_pasivo = {
        "fecha": as_of, "activo": activo, "pasivo": pasivo,
        "base_monetaria": bm, "instrumentos": (pases or 0) + (letras or 0),
        "reservas_usd": res_usd, "tipo_cambio": fx,
        "cobertura_pct": (activo / pasivo * 100) if pasivo else None,
    }

# ---------- 5. armar informe.html (portada del sitio) ----------
def fmt_ar(n, decimals=2):
    if n is None:
        return "s/d"
    s = f"{n:,.{decimals}f}"
    return s.replace(",", "§").replace(".", ",").replace("§", ".")

def fmt_date_ddmmyyyy(iso):
    if not iso:
        return "s/d"
    y, m, d = iso.split("-")
    return f"{d}/{m}/{y}"

ap_html = ""
if activo_pasivo:
    ap = activo_pasivo
    max_total = max(ap["activo"], ap["pasivo"]) or 1
    w_activo = ap["activo"] / max_total * 100
    w_bm = ap["base_monetaria"] / max_total * 100
    w_ins = ap["instrumentos"] / max_total * 100
    cobertura_txt = f'<div class="ap-cov">Las reservas cubren <b>{fmt_ar(ap["cobertura_pct"], 0)}%</b> del pasivo</div>' if ap["cobertura_pct"] is not None else ""
    ap_html = f"""
    <div class="ap-card">
      <div class="ap-head">
        <div class="section-title" style="margin:0;font-size:14px;">Balance simplificado del BCRA</div>
        {cobertura_txt}
      </div>
      <div class="ap-row">
        <div class="ap-row-label"><span>Activo — Reservas internacionales</span><span class="ap-row-total">${fmt_ar(ap['activo']/1e6)} Bn</span></div>
        <div class="ap-bar-track"><div class="ap-bar-fill activo" style="width:{w_activo:.1f}%"></div></div>
      </div>
      <div class="ap-row">
        <div class="ap-row-label"><span>Pasivo — Base monetaria + instrumentos remunerados</span><span class="ap-row-total">${fmt_ar(ap['pasivo']/1e6)} Bn</span></div>
        <div class="ap-bar-track">
          <div class="ap-bar-fill bm" style="width:{w_bm:.1f}%"></div>
          <div class="ap-bar-fill ins" style="width:{w_ins:.1f}%"></div>
        </div>
      </div>
      <div class="ap-legend">
        <span><i class="dot activo"></i>Reservas (US$ {fmt_ar(ap['reservas_usd'],0)} M a ${fmt_ar(ap['tipo_cambio'])})</span>
        <span><i class="dot bm"></i>Base monetaria</span>
        <span><i class="dot ins"></i>Pases + letras/notas BCRA</span>
      </div>
      <p class="section-sub" style="margin:10px 0 0;font-size:11px;">Aproximación: no descuenta pasivos en
        dólares del BCRA ni otras partidas del balance real. Al {fmt_date_ddmmyyyy(ap['fecha'])}.</p>
    </div>"""

TIPO_LABELS_UI = {"bono": "Bonos", "letra": "Letras", "otra_operacion": "Otras operaciones"}
lic_cards_html = ""
for lic in licitaciones[:8]:
    re_ = lic.get("resumenExcel") or {}
    ve_ars = re_.get("veAdjudicadoArs") or 0
    ve_usd = re_.get("veAdjudicadoUsd") or 0
    n_instr = re_.get("cantidadInstrumentos", 0)
    ars_txt = f"${fmt_ar(ve_ars/1e6)} Bn" if ve_ars else "s/d"
    usd_txt = f' · USD {fmt_ar(ve_usd,0)} M' if ve_usd else ""
    comp = lic.get("composicion") or []
    comp_txt = " · ".join(f"{c['categoria']} {fmt_ar((c.get('pct') or 0)*100,0)}%" for c in comp if c.get("categoria") != "TOTAL")
    lic_cards_html += f"""
      <div class="lic-item">
        <div class="lic-item-date">{fmt_date_ddmmyyyy(lic.get('fecha'))}</div>
        <div class="lic-item-val">{ars_txt}{usd_txt}</div>
        <div class="lic-item-sub">{n_instr} instrumento{'s' if n_instr != 1 else ''} colocado{'s' if n_instr != 1 else ''}{' · ' + comp_txt if comp_txt else ''}</div>
      </div>"""

INFORME_HTML = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Licitaciones e Indicadores Monetarios - Argentina</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
  :root{{
    --navy:#0B2547; --blue:#2C5FA8; --blue-light:#E7EFFB; --sky:#5B8FD9;
    --green:#1E8A5F; --amber:#C2571B; --paper:#F6F7FB; --ink:#0B2547; --ink-soft:#5B6472; --line:#DCE3EF;
  }}
  *{{box-sizing:border-box;}}
  html,body{{margin:0;background:#DCE3EF;font-family:'Inter',sans-serif;color:var(--ink);}}
  body{{display:flex;justify-content:center;padding:32px 12px;}}
  .page{{width:900px;max-width:100%;background:var(--paper);border-radius:22px;overflow:hidden;box-shadow:0 30px 60px -20px rgba(11,37,71,.35);}}
  .hero{{background:linear-gradient(135deg,var(--navy) 0%,#173B72 60%,var(--blue) 100%);color:#fff;padding:36px 44px 30px;position:relative;overflow:hidden;}}
  .hero::after{{content:"";position:absolute;right:-60px;top:-60px;width:260px;height:260px;border-radius:50%;background:rgba(255,255,255,.06);}}
  .eyebrow{{font-size:12.5px;letter-spacing:.14em;text-transform:uppercase;color:#AFC7EE;font-weight:600;margin-bottom:10px;}}
  h1{{font-family:'Space Grotesk',sans-serif;font-size:26px;line-height:1.2;margin:0 0 8px;font-weight:700;max-width:620px;}}
  .hero p{{margin:0;color:#D7E3F6;font-size:14px;max-width:600px;}}
  .hero a{{color:#AFC7EE;}}
  .section{{padding:26px 44px 4px;}}
  .section-title{{font-family:'Space Grotesk',sans-serif;font-size:16px;font-weight:700;color:var(--navy);margin:0 0 4px;}}
  .section-sub{{font-size:12.5px;color:var(--ink-soft);margin:0 0 14px;line-height:1.5;}}
  .ap-card{{background:#fff;border:1px solid var(--line);border-radius:14px;padding:18px 22px;margin:0 44px 22px;}}
  .ap-head{{display:flex;justify-content:space-between;align-items:baseline;gap:10px;flex-wrap:wrap;margin-bottom:14px;}}
  .ap-cov{{font-size:12px;color:var(--ink-soft);}}
  .ap-cov b{{color:var(--navy);font-family:'Space Grotesk',sans-serif;}}
  .ap-row{{margin-bottom:10px;}}
  .ap-row-label{{display:flex;justify-content:space-between;font-size:12px;color:var(--ink-soft);margin-bottom:4px;}}
  .ap-row-total{{font-family:'Space Grotesk',sans-serif;font-weight:700;color:var(--navy);}}
  .ap-bar-track{{display:flex;height:16px;border-radius:8px;overflow:hidden;background:var(--line);}}
  .ap-bar-fill{{height:100%;}}
  .ap-bar-fill.activo{{background:var(--green);}}
  .ap-bar-fill.bm{{background:var(--blue);}}
  .ap-bar-fill.ins{{background:var(--amber);}}
  .ap-legend{{display:flex;gap:16px;flex-wrap:wrap;font-size:11px;color:var(--ink-soft);margin-top:10px;}}
  .ap-legend .dot{{display:inline-block;width:8px;height:8px;border-radius:50%;margin-right:5px;}}
  .ap-legend .dot.activo{{background:var(--green);}} .ap-legend .dot.bm{{background:var(--blue);}} .ap-legend .dot.ins{{background:var(--amber);}}
  .lic-list{{display:grid;gap:10px;margin:0 44px 8px;}}
  .lic-item{{background:#fff;border:1px solid var(--line);border-radius:12px;padding:12px 16px;display:flex;justify-content:space-between;align-items:center;gap:14px;flex-wrap:wrap;}}
  .lic-item-date{{font-family:'Space Grotesk',sans-serif;font-weight:700;color:var(--navy);font-size:13px;min-width:80px;}}
  .lic-item-val{{font-family:'Space Grotesk',sans-serif;font-weight:700;color:var(--blue);font-size:13px;}}
  .lic-item-sub{{font-size:11.5px;color:var(--ink-soft);flex:1;text-align:right;}}
  .footer{{padding:20px 44px 30px;font-size:11px;color:#8A93A3;line-height:1.6;}}
</style>
</head>
<body>
<div class="page">
  <div class="hero">
    <div class="eyebrow">Coyuntura monetaria argentina</div>
    <h1>Licitaciones del Tesoro e indicadores monetarios del BCRA</h1>
    <p>Actualizado automáticamente a partir de dos fuentes oficiales: el registro de Colocaciones de Deuda de
      la Oficina Nacional de Crédito Público y las series monetarias diarias del BCRA.</p>
    <p style="margin-top:12px;"><a href="monetario.html">Ir al explorador interactivo →</a></p>
  </div>
  {ap_html}
  <div class="section">
    <div class="section-title">Últimas licitaciones</div>
    <div class="section-sub">Valor efectivo adjudicado por fecha de colocación. Ver el detalle completo, con
      composición e instrumentos, en el explorador.</div>
  </div>
  <div class="lic-list">{lic_cards_html or '<p class="section-sub">Todavía no hay licitaciones archivadas.</p>'}</div>
  <div class="footer">
    Fuentes: registro de Colocaciones de Deuda (Oficina Nacional de Crédito Público) y series.xlsm
    (Gerencia de Estadísticas Monetarias, BCRA — datos provisorios sujetos a revisión). Página generada
    automáticamente.
  </div>
</div>
</body>
</html>
"""

out_informe = HERE / "informe.html"
out_informe.write_text(INFORME_HTML, encoding="utf-8")
print(f"HTML generado: {out_informe}")
