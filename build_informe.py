#!/usr/bin/env python3
"""
Genera informe.html + informe.pdf a partir de datos_informe.xlsx
Uso: python3 build_informe.py [ruta_excel]
"""
import sys, subprocess
import openpyxl
from pathlib import Path

HERE = Path(__file__).parent
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else HERE / "datos_informe.xlsx"

wb = openpyxl.load_workbook(XLSX, data_only=True)

def sheet_rows(name, start=2):
    ws = wb[name]
    for row in ws.iter_rows(min_row=start, values_only=True):
        if row[0] is None:
            continue
        yield row

def kv(name):
    return {r[0]: r[1] for r in sheet_rows(name)}

portada = {r[0]: r[1] for r in wb["Portada"].iter_rows(min_row=1, values_only=True) if r[0]}
rollover = kv("Rollover")
tasas = list(sheet_rows("Tasas"))
coyuntura = list(sheet_rows("Coyuntura"))
composicion = [r for r in sheet_rows("Composicion") if r[3]]  # descarta fila TOTAL (sin color)
timeline = list(sheet_rows("Timeline"))
monet = [r for r in sheet_rows("Monetizacion") if r[2]]
fuente_txt = wb["Fuentes"]["A2"].value

# ---------- helpers ----------
from datetime import date, datetime
def parse_d(d):
    if not d:
        return None
    if isinstance(d, (date, datetime)):
        return d
    return datetime.strptime(str(d), "%Y-%m-%d").date()

MESES = {1:"ene",2:"feb",3:"mar",4:"abr",5:"may",6:"jun",7:"jul",8:"ago",9:"sep",10:"oct",11:"nov",12:"dic"}
def fmt_date(d):
    d = parse_d(d)
    if not d:
        return None
    return f"{MESES[d.month]}-{str(d.year)[2:]}"

def month_index(d, base_year=2026, base_month=7):
    d = parse_d(d)
    return (d.year - base_year) * 12 + (d.month - base_month)

TL_MIN, TL_MAX = 0, 39  # jul-2026 .. oct-2029
def x_pct(d):
    mi = month_index(d)
    return max(0, min(100, (mi - TL_MIN) / (TL_MAX - TL_MIN) * 100))

TAG_CLASS = {"warn":"warn","good":"good","flat":"flat"}
MOVE_CLASS = {"▲":"up","▼":"down","—":"flat"}

def move_class(text):
    for sym,c in MOVE_CLASS.items():
        if text.startswith(sym):
            return c
    return "flat"

# ---------- KPI cards on hero (rate cards) ----------
rate_cards_html = ""
for instr, now, before, direction in tasas:
    pill = "up" if direction == "subio" else "down"
    arrow = "▲" if direction=="subio" else "▼"
    label = "subió" if direction=="subio" else "bajó"
    rate_cards_html += f"""
    <div class="rate-card">
      <div class="name">{instr}</div>
      <div class="now">{now:.2f}%</div>
      <div class="then">antes: {before:.2f}%</div>
      <span class="pill {pill}">{arrow} {label}</span>
    </div>"""

# ---------- coyuntura rows ----------
coy_html = ""
for instr, plazo, tasa, lectura, tipo in coyuntura:
    coy_html += f"""
    <div class="coy-row">
      <span class="instr">{instr}</span>
      <span class="coy-move {move_class(plazo)}">{plazo}</span>
      <span class="coy-move {move_class(tasa)}">{tasa}</span>
      <span class="coy-tag {TAG_CLASS.get(tipo,'flat')}">{lectura}</span>
    </div>"""

# ---------- composicion (donut) ----------
total = sum(r[1] for r in composicion)
circ_r, circ_w = 72, 34
circumf = 2 * 3.14159265 * circ_r
donut_circles = ""
legend_html = ""
offset = 0.0
for cat, monto, pct, color in composicion:
    frac = monto / total
    length = circumf * frac
    donut_circles += (f'<circle cx="100" cy="100" r="{circ_r}" fill="none" stroke="#{color}" '
                       f'stroke-width="{circ_w}" stroke-dasharray="{length:.1f} {circumf-length:.1f}" '
                       f'stroke-dashoffset="-{offset:.1f}" transform="rotate(-90 100 100)"/>\n')
    legend_html += (f'<div class="legend-item"><span class="legend-dot" style="background:#{color}"></span>'
                     f'{cat} <b>{round(frac*100)}%</b></div>\n')
    offset += length

pesos_var = round(sum(r[1] for r in composicion if r[3] in ("5B8FD9","1E8A5F")) / total * 100)
fija_pct = round([r[1] for r in composicion if r[0]=="Pesos, tasa fija"][0] / total * 100)
composicion_note = (f"Más de la mitad de lo colocado ({pesos_var}%) quedó atado al dólar o a tasa variable "
                     f"(dual TAMAR/DL + dólar linked puro), contra apenas {fija_pct}% a tasa fija pura en pesos.")

# ---------- timeline ----------
tl_html = ""
for instr, d_antes, d_ahora, caption, color in timeline:
    dots = ""
    if d_antes:
        dots += f'<div class="tl-dot" style="left:{x_pct(d_antes):.1f}%; background:#B4B2A9;"></div>'
    dots += f'<div class="tl-dot" style="left:{x_pct(d_ahora):.1f}%; background:#{color};"></div>'
    tl_html += f"""
    <div class="tl-row">
      <div class="tl-label-row">
        <span class="tl-label">{instr}</span>
        <span class="tl-caption">{caption}</span>
      </div>
      <div class="tl-track">{dots}</div>
    </div>"""

# ---------- monetizacion ----------
m = {r[0]: (r[1], r[2]) for r in monet}
monet_html = ""
for concepto, (valor, unidad) in m.items():
    if "%" in unidad:
        val_str = f"{valor:.1f}%"
    else:
        val_str = f"${valor:.2f} Bn" if "billones" in unidad else f"{valor}"
    monet_html += f"""
    <div class="mon-row">
      <span class="mon-concept">{concepto}</span>
      <span class="mon-val">{val_str}</span>
    </div>"""

# ---------- rollover gauge ----------
pct = rollover["Rollover (%)"]
gauge_frac = min(pct/100, 1.5) / 1.5  # cap visual fill at 150%
gauge_circumf = 2 * 3.14159265 * 90

HTML = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>{portada.get('Título','Informe')}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
  :root{{
    --navy:#0B2547; --blue:#2C5FA8; --blue-light:#E7EFFB; --sky:#5B8FD9;
    --green:#1E8A5F; --green-light:#E6F5EE; --amber:#C2571B; --amber-light:#FCEEE3;
    --paper:#F6F7FB; --ink:#0B2547; --ink-soft:#5B6472; --line:#DCE3EF;
  }}
  *{{box-sizing:border-box;}}
  body{{margin:0;background:#DCE3EF;font-family:'Inter',sans-serif;color:var(--ink);display:flex;justify-content:center;padding:32px 12px;}}
  .page{{width:900px;max-width:100%;background:var(--paper);border-radius:22px;overflow:hidden;box-shadow:0 30px 60px -20px rgba(11,37,71,.35);}}
  .hero{{background:linear-gradient(135deg,var(--navy) 0%,#173B72 60%,var(--blue) 100%);color:#fff;padding:36px 44px 30px;position:relative;overflow:hidden;}}
  .hero::after{{content:"";position:absolute;right:-60px;top:-60px;width:260px;height:260px;border-radius:50%;background:rgba(255,255,255,.06);}}
  .eyebrow{{font-size:12.5px;letter-spacing:.14em;text-transform:uppercase;color:#AFC7EE;font-weight:600;margin-bottom:10px;}}
  h1{{font-family:'Space Grotesk',sans-serif;font-size:28px;line-height:1.15;margin:0 0 8px;font-weight:700;max-width:600px;}}
  .hero p{{margin:0;color:#D7E3F6;font-size:14.5px;max-width:580px;}}
  .verdict{{display:grid;grid-template-columns:1fr 1fr;background:#fff;}}
  .verdict > div{{padding:26px 40px;display:flex;gap:16px;align-items:flex-start;}}
  .verdict > div:first-child{{border-right:1px solid var(--line);}}
  .badge{{flex:none;width:44px;height:44px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:20px;font-weight:700;color:#fff;}}
  .badge.ok{{background:var(--green);}} .badge.warn{{background:var(--amber);}}
  .verdict h3{{font-family:'Space Grotesk',sans-serif;font-size:14px;margin:2px 0 4px;text-transform:uppercase;letter-spacing:.04em;}}
  .verdict h3.ok-text{{color:var(--green);}} .verdict h3.warn-text{{color:var(--amber);}}
  .verdict p{{margin:0;font-size:14px;color:var(--ink-soft);line-height:1.45;}} .verdict b{{color:var(--ink);}}
  .section{{padding:30px 44px 8px;}}
  .section-label{{font-family:'Space Grotesk',sans-serif;font-size:12.5px;letter-spacing:.1em;text-transform:uppercase;color:var(--blue);font-weight:700;margin-bottom:2px;}}
  .section-title{{font-family:'Space Grotesk',sans-serif;font-size:19px;font-weight:600;color:var(--navy);margin:0 0 16px;}}
  .rollover-wrap{{padding:0 44px 8px;display:grid;grid-template-columns:230px 1fr;gap:32px;align-items:center;}}
  .gauge{{position:relative;width:210px;height:210px;}}
  .gauge svg{{transform:rotate(-90deg);}}
  .gauge-center{{position:absolute;inset:0;display:flex;flex-direction:column;align-items:center;justify-content:center;}}
  .gauge-center .num{{font-family:'Space Grotesk',sans-serif;font-size:34px;font-weight:700;color:var(--navy);}}
  .gauge-center .lbl{{font-size:11px;color:var(--ink-soft);text-transform:uppercase;letter-spacing:.06em;margin-top:2px;}}
  .rollover-explain p{{font-size:14.5px;line-height:1.6;color:var(--ink-soft);margin:0 0 14px;}}
  .rollover-explain b{{color:var(--navy);}}
  .bars{{display:flex;flex-direction:column;gap:10px;}}
  .bar-row{{display:grid;grid-template-columns:150px 1fr 70px;align-items:center;gap:10px;font-size:13px;}}
  .bar-track{{background:var(--line);border-radius:8px;height:14px;overflow:hidden;}}
  .bar-fill{{height:100%;border-radius:8px;}}
  .bar-fill.vencimientos{{background:var(--sky);}} .bar-fill.adjudicado{{background:var(--green);}}
  .rates{{padding:6px 44px 10px;display:grid;grid-template-columns:repeat(4,1fr);gap:14px;}}
  .rate-card{{background:#fff;border:1px solid var(--line);border-radius:14px;padding:16px 16px 14px;}}
  .rate-card .name{{font-size:12px;color:var(--ink-soft);margin-bottom:10px;min-height:30px;line-height:1.3;}}
  .rate-card .now{{font-family:'Space Grotesk',sans-serif;font-size:24px;font-weight:700;color:var(--navy);}}
  .rate-card .then{{font-size:11.5px;color:var(--ink-soft);margin-top:2px;}}
  .pill{{display:inline-flex;align-items:center;gap:4px;font-size:11.5px;font-weight:700;padding:2px 8px;border-radius:20px;margin-top:8px;}}
  .pill.up{{background:var(--amber-light);color:var(--amber);}} .pill.down{{background:var(--green-light);color:var(--green);}}
  .coy-row{{display:grid;grid-template-columns:1.5fr 1fr 1fr 1.4fr;align-items:center;gap:10px;background:#fff;border:1px solid var(--line);border-radius:12px;padding:12px 16px;margin-bottom:8px;}}
  .coy-row .instr{{font-size:13px;font-weight:600;color:var(--navy);}}
  .coy-move{{font-size:12.5px;display:flex;align-items:center;gap:4px;}}
  .coy-move.up{{color:var(--amber);}} .coy-move.down{{color:var(--green);}} .coy-move.flat{{color:var(--ink-soft);}}
  .coy-tag{{font-size:11.5px;font-weight:600;text-align:center;padding:5px 10px;border-radius:20px;}}
  .coy-tag.warn{{background:var(--amber-light);color:var(--amber);}}
  .coy-tag.good{{background:var(--green-light);color:var(--green);}}
  .coy-tag.flat{{background:#EEF0F4;color:var(--ink-soft);}}
  .donut-wrap{{padding:0 44px 6px;display:grid;grid-template-columns:200px 1fr;gap:32px;align-items:center;}}
  .legend-list{{display:flex;flex-direction:column;gap:10px;}}
  .legend-item{{display:flex;align-items:center;gap:10px;font-size:13.5px;}}
  .legend-dot{{width:12px;height:12px;border-radius:3px;flex:none;}}
  .legend-item b{{color:var(--navy);margin-left:auto;font-family:'Space Grotesk',sans-serif;}}
  .tl-row{{margin-bottom:20px;}}
  .tl-label-row{{display:flex;justify-content:space-between;align-items:baseline;margin-bottom:6px;}}
  .tl-label{{font-size:13px;font-weight:600;color:var(--navy);}}
  .tl-caption{{font-size:11.5px;color:var(--ink-soft);}}
  .tl-track{{position:relative;height:10px;background:var(--line);border-radius:6px;}}
  .tl-dot{{position:absolute;top:50%;transform:translate(-50%,-50%);width:15px;height:15px;border-radius:50%;border:2px solid #fff;}}
  .tl-axis{{display:flex;justify-content:space-between;margin:0 0 10px;padding-top:4px;border-top:1px solid var(--line);}}
  .tl-axis span{{font-size:10.5px;color:var(--ink-soft);font-weight:600;}}
  .mon-row{{display:flex;justify-content:space-between;align-items:center;background:#fff;border:1px solid var(--line);border-radius:12px;padding:12px 16px;margin-bottom:8px;font-size:13px;}}
  .mon-concept{{color:var(--navy);}}
  .mon-val{{font-family:'Space Grotesk',sans-serif;font-weight:700;color:var(--blue);}}
  .footer{{padding:20px 44px 30px;font-size:11px;color:#8A93A3;line-height:1.6;}}
</style>
</head>
<body>
<div class="page">

  <div class="hero">
    <div class="eyebrow">{portada.get('Eyebrow','')}</div>
    <h1>{portada.get('Título','')}</h1>
    <p>{portada.get('Subtítulo','')}</p>
    <p style="margin-top:12px;"><a href="monetario.html" style="color:#AFC7EE;">Ver indicadores monetarios (BCRA) →</a></p>
  </div>

  <div class="verdict">
    <div>
      <div class="badge ok">✓</div>
      <div>
        <h3 class="ok-text">{portada.get('Veredicto 1 - título','')}</h3>
        <p>{portada.get('Veredicto 1 - texto (usar {venc},{adj},{pct},{extra} como placeholders)','').format(venc=rollover['Vencimientos del día ($ Bn)'], adj=rollover['Total adjudicado ($ Bn)'], pct=rollover['Rollover (%)'], extra=rollover['Excedente absorbido ($ Bn)'])}</p>
      </div>
    </div>
    <div>
      <div class="badge warn">!</div>
      <div>
        <h3 class="warn-text">{portada.get('Veredicto 2 - título','')}</h3>
        <p>{portada.get('Veredicto 2 - texto','')}</p>
      </div>
    </div>
  </div>

  <div class="section"><div class="section-label">Parte 1</div><div class="section-title">¿Alcanzó para cubrir lo que vencía?</div></div>
  <div class="rollover-wrap">
    <div class="gauge">
      <svg width="210" height="210" viewBox="0 0 210 210">
        <circle cx="105" cy="105" r="90" fill="none" stroke="#DCE3EF" stroke-width="20"/>
        <circle cx="105" cy="105" r="90" fill="none" stroke="#1E8A5F" stroke-width="20"
                stroke-dasharray="{gauge_circumf*gauge_frac:.1f} {gauge_circumf:.1f}" stroke-dashoffset="0" stroke-linecap="round"/>
      </svg>
      <div class="gauge-center"><div class="num">{pct:.1f}%</div><div class="lbl">Rollover</div></div>
    </div>
    <div class="rollover-explain">
      <p>De cada <b>$100</b> que el Tesoro tenía que pagar, consiguió <b>${pct:.1f}</b> — cubrió todo lo que vencía y absorbió plata extra del mercado.</p>
      <div class="bars">
        <div class="bar-row"><span>Vencía</span><div class="bar-track"><div class="bar-fill vencimientos" style="width:{rollover['Vencimientos del día ($ Bn)']/rollover['Total adjudicado ($ Bn)']*100:.0f}%"></div></div><span>${rollover['Vencimientos del día ($ Bn)']:.2f} Bn</span></div>
        <div class="bar-row"><span>Consiguió</span><div class="bar-track"><div class="bar-fill adjudicado" style="width:100%"></div></div><span>${rollover['Total adjudicado ($ Bn)']:.2f} Bn</span></div>
      </div>
    </div>
  </div>

  <div class="section" style="padding-top:26px;"><div class="section-label">Parte 2</div><div class="section-title">¿Pagó más o menos tasa que la vez anterior?</div></div>
  <div class="rates">{rate_cards_html}</div>

  <div class="section" style="padding-top:26px;"><div class="section-label">Parte 3</div><div class="section-title">Lectura: ¿plazo y tasa mejoraron juntos o no?</div></div>
  <div class="section" style="padding-top:0;padding-bottom:6px;">{coy_html}</div>
  <div class="section" style="padding-top:26px;"><div class="section-label">Parte 4</div><div class="section-title">¿En qué puso la plata el Tesoro?</div></div>
  <div class="donut-wrap">
    <svg width="200" height="200" viewBox="0 0 200 200">
      <circle cx="100" cy="100" r="72" fill="none" stroke="#EEF0F4" stroke-width="34"/>
      {donut_circles}
    </svg>
    <div class="legend-list">{legend_html}</div>
  </div>
  <div class="section" style="padding-top:2px;padding-bottom:0;">
    <p style="font-size:13.5px;color:var(--ink-soft);line-height:1.55;margin:6px 0 0;">{composicion_note}</p>
  </div>

  <div class="section" style="padding-top:26px;"><div class="section-label">Parte 5</div><div class="section-title">¿Cuándo vence ahora cada cosa?</div></div>
  <div class="section" style="padding-top:0;padding-bottom:10px;">
    <div class="tl-axis"><span>jul-2026</span><span>2027</span><span>2028</span><span>2029</span></div>
    {tl_html}
  </div>

  <div class="section" style="padding-top:26px;"><div class="section-label">Parte 6</div><div class="section-title">¿Cómo queda la cantidad de pesos en la economía?</div></div>
  <div class="section" style="padding-top:0;padding-bottom:6px;">{monet_html}</div>
  <div class="footer">{fuente_txt}</div>

</div>
</body>
</html>
"""

out_html = HERE / "informe.html"
out_html.write_text(HTML, encoding="utf-8")
print(f"HTML generado: {out_html}")

# ---------- Render a PDF con Playwright ----------
from playwright.sync_api import sync_playwright
with sync_playwright() as p:
    b = p.chromium.launch()
    page = b.new_page(viewport={"width": 960, "height": 1200})
    page.goto(f"file://{out_html}")
    page.wait_for_timeout(300)
    height = page.evaluate("document.body.scrollHeight")
    out_pdf = HERE / "informe.pdf"
    page.pdf(path=str(out_pdf), width="960px", height=f"{height+40}px",
             print_background=True, margin={"top":"20px","bottom":"20px","left":"20px","right":"20px"})
    b.close()
print(f"PDF generado: {out_pdf}")
